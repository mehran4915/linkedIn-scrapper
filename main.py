# Mehran Ahmadzadeh
# Ehsan rasouli

import json
from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from webdriver_manager.firefox import GeckoDriverManager
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import time
# env
import os
from dotenv import load_dotenv
# excel
from openpyxl import Workbook

class LinkedInJobScraper:
    def __init__(self, job_title, location):
        self.job_title = job_title
        self.location = location
        self.jobs_data = []

        options = Options()

        self.driver = webdriver.Firefox(
            service=Service(GeckoDriverManager().install()),
        )
        self.driver.get("https://www.linkedin.com")
        time.sleep(3)

    def load_cookies(self):
        load_dotenv()
        cookies_str = os.getenv("LINKEDIN_COOKIES")
        cookies = json.loads(cookies_str)

        for cookie in cookies:
            try:
                self.driver.add_cookie(cookie)
            except Exception as e:
                print(f"Skipping cookie {cookie.get('name')}: {e}")


        self.driver.refresh()
        time.sleep(5)

        self.driver.get("https://www.linkedin.com/jobs/search")
        time.sleep(5)

    def scrape_jobs(self, number_of_jobs_needed):
        self.load_cookies()

        #--------- entring job and location -----------#

        job_input = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'input[aria-label="Search by title, skill, or company"]:not([disabled])')
            )
        )
        job_input.click()
        time.sleep(1)  
        job_input.send_keys(Keys.CONTROL + "a")  
        job_input.send_keys(Keys.BACKSPACE)     
        job_input.send_keys(self.job_title)

        location_input = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, 'input[aria-label="City, state, or zip code"]:not([disabled])')
            )
        )
        location_input.clear()
        location_input.send_keys(self.location)

        search_button = WebDriverWait(self.driver, 10).until(
            EC.element_to_be_clickable(
                (By.CSS_SELECTOR, "button.jobs-search-box__submit-button")
            )
        )
        search_button.click()
        time.sleep(5)

        # ------ find the jobs ----- #

        job_items = WebDriverWait(self.driver, 10).until(
            EC.presence_of_all_elements_located((By.CSS_SELECTOR, "li[data-occludable-job-id]"))
        )

        for li in job_items[:number_of_jobs_needed]:
            try:
                job_id = li.get_attribute("data-occludable-job-id")

                self.driver.execute_script("arguments[0].scrollIntoView({block:'center'});", li)

                title_elem = WebDriverWait(self.driver, 6).until(
                    EC.presence_of_element_located(
                        (By.CSS_SELECTOR, f"li[data-occludable-job-id='{job_id}'] a[aria-label][href*='/jobs/view/']")
                    )
                )
                job_title = title_elem.get_attribute("aria-label").strip()
                job_link = title_elem.get_attribute("href").strip()

                try:
                    company_elem = self.driver.find_element(
                        By.CSS_SELECTOR,
                        f"li[data-occludable-job-id='{job_id}'] .artdeco-entity-lockup__subtitle span"
                    )
                    company_name = company_elem.text.strip()
                except:
                    company_name = ""

                try:
                    location_elem = self.driver.find_element(
                        By.CSS_SELECTOR,
                        f"li[data-occludable-job-id='{job_id}'] .artdeco-entity-lockup__caption .job-card-container__metadata-wrapper li span"
                    )
                    job_location = location_elem.text.strip()
                except:
                    job_location = ""

                try:
                    salary_elem = self.driver.find_element(
                        By.CSS_SELECTOR,
                        f"li[data-occludable-job-id='{job_id}'] .artdeco-entity-lockup__metadata .job-card-container__metadata-wrapper li span"
                    )
                    salary = salary_elem.text.strip()
                except:
                    salary = ""

                self.jobs_data.append({
                    "title": job_title,
                    "company": company_name,
                    "location": job_location,
                    "salary": salary,
                    "link": job_link
                })

            except Exception as e:
                print(f"Skipping job {job_id}: {e}")
        
        self.driver.quit()

    def save_to_excel(self, filename):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Jobs"

            # Head
            headers = ["Title", "Company", "Location", "Salary", "Link"]
            ws.append(headers)

            # rows
            for job in self.jobs_data:
                row = [
                    job["title"],
                    job["company"],
                    job["location"],
                    job["salary"],
                    job["link"]
                ]
                ws.append(row)
                ws.cell(ws.max_row, 5).hyperlink = job["link"]
                ws.cell(ws.max_row, 5).style = "Hyperlink"

            for col in ws.columns:
                if col[0].column_letter == "C":
                    ws.column_dimensions[col[0].column_letter].width = 50
                else:
                    ws.column_dimensions[col[0].column_letter].width = 25

            wb.save(filename)
            print(f"✅ Jobs saved to {filename}")

        except Exception as e:
            print(f"Failed to save Excel: {e}")



if __name__ == "__main__":
    job_wanted = input("what job are you looking for ?  ")
    location_wanted = input("Where do you want to work ?  ")
    scraper = LinkedInJobScraper( job_wanted , location_wanted)
    scraper.scrape_jobs(4)
    scraper.save_to_excel(f"{job_wanted}_{location_wanted}.xlsx")
